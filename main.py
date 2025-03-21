import requests
import camelot.io as camelot
import openpyxl

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
import config
import time

from fbchat import Client
from fbchat.models import *

# Geting the PDF returning an error if can not
try:
    print("Downloading PDF...")
    URL = "https://vasvari.org/wp-content/uploads/2022-okt_-17.pdf"
    response = requests.get(URL)
    open("files/file.pdf", "wb").write(response.content)
    print("PDF downloaded")
except:
    print("An error has accourd when downloading and convertin the PDF file.")

# Convert PDF tabel to .xlsx file
file = "files/file.pdf"
tables = camelot.read_pdf(file)
tables[0].to_excel("files/output.xlsx")

# Opening the .xlsx file an deletin unnesesary row and culomns
path = "files/output.xlsx"
wb = openpyxl.load_workbook(path)
sheet = wb.active

row = sheet.max_row
column = sheet.max_column

sheet.move_range("B4:L" + str(row + 3), rows=-3, cols=-1)

if (sheet['C3'].value == None):
    print("Deleting unnessesary rows...")
    sheet.delete_rows(3)

wb.save("files/format.xlsx")

print("Finished")

row = sheet.max_row
classes = []
#Search class
for i in len(range(row)):
    if("11.c" in sheet['C'+i].value):
        classes.append(i)

driver = webdriver.Chrome("../chromedriver.exe")

# driver.add_cookie({"name": "messenger.com", "value": "WWdGY4wZuqNx-uigp68IN6wb"})
# driver.add_cookie({"domain": ".messenger.com", "expiry'": 1700118373, "httpOnly": True, "name": "datr", "path": "/", "sameSite": None, "secure": True, "value": "WWdGY4wZuqNx-uigp68IN6wb"})


driver.get("https://www.messenger.com/t/2022213241239792")

print("Current email is '" + config.email + "' Current password is '" + config.password + "'")

time.sleep(7)

print("Login in...")
driver.find_element(By.NAME, "email").send_keys(config.email)
driver.find_element(By.NAME, "pass").send_keys(config.password + Keys.RETURN)

time.sleep(5)

#print(driver.find_element(By.ID, "jsc_c_3p"))
driver.find_element(By.XPATH, "/html/body/div[1]/div/div/div/div[2]/div/div/div/div[1]/div[1]/div[2]/div/div/div/div/div/div[2]/div/div/div[2]/div/div/div[4]/div[2]/div/div/div[1]/p").send_keys(
   "Halal a feketerkre (Content generated by adri.py)" + Keys.RETURN)

time.sleep(10)

driver.quit()





# CREATE TABLE all_date (
#	 type varchar(4),
#    lesson varchar(3),
#    class varchar(10),
#    inTheTimetabel varchar(10),
#    actualLesson varchar(10),
#    substituteTeacher varchar(20),
#    lessonDate varchar(10),
#    lessonNumber varchar(5),
#    fromWhere varchar(6),
#    toWhere varchar(6)
# )
